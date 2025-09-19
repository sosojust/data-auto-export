import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo
from datetime import datetime
from loguru import logger
from typing import Optional, Dict, Any, List

class ExcelExporter:
    """Excel导出器"""
    
    def __init__(self, output_dir: str = "./exports"):
        self.output_dir = output_dir
        os.makedirs(output_dir, exist_ok=True)
    
    def export_dataframe(self, 
                        df: pd.DataFrame, 
                        filename: str,
                        sheet_name: str = "Sheet1",
                        apply_formatting: bool = True,
                        add_table_style: bool = True) -> str:
        """导出DataFrame到Excel文件"""
        
        # 确保文件名以.xlsx结尾
        if not filename.endswith('.xlsx'):
            filename += '.xlsx'
        
        file_path = os.path.join(self.output_dir, filename)
        
        try:
            # 创建工作簿
            wb = Workbook()
            ws = wb.active
            ws.title = sheet_name
            
            # 如果DataFrame为空，创建一个提示
            if df.empty:
                ws['A1'] = "查询结果为空"
                ws['A1'].font = Font(bold=True, color="FF0000")
                wb.save(file_path)
                logger.warning(f"导出的DataFrame为空: {file_path}")
                return file_path
            
            # 将DataFrame写入工作表
            for r in dataframe_to_rows(df, index=False, header=True):
                ws.append(r)
            
            # 应用格式化
            if apply_formatting:
                self._apply_formatting(ws, df)
            
            # 添加表格样式
            if add_table_style and len(df) > 0:
                self._add_table_style(ws, df)
            
            # 自动调整列宽
            self._auto_adjust_column_width(ws)
            
            # 保存文件
            wb.save(file_path)
            
            file_size = os.path.getsize(file_path)
            logger.info(f"Excel文件导出成功: {file_path}, 大小: {file_size} bytes, 行数: {len(df)}")
            
            return file_path
            
        except Exception as e:
            logger.error(f"Excel导出失败: {e}")
            raise
    
    def export_multiple_sheets(self, 
                              data_dict: Dict[str, pd.DataFrame], 
                              filename: str,
                              apply_formatting: bool = True) -> str:
        """导出多个工作表到一个Excel文件"""
        
        if not filename.endswith('.xlsx'):
            filename += '.xlsx'
        
        file_path = os.path.join(self.output_dir, filename)
        
        try:
            wb = Workbook()
            # 删除默认工作表
            wb.remove(wb.active)
            
            for sheet_name, df in data_dict.items():
                ws = wb.create_sheet(title=sheet_name)
                
                if df.empty:
                    ws['A1'] = "查询结果为空"
                    ws['A1'].font = Font(bold=True, color="FF0000")
                    continue
                
                # 写入数据
                for r in dataframe_to_rows(df, index=False, header=True):
                    ws.append(r)
                
                # 应用格式化
                if apply_formatting:
                    self._apply_formatting(ws, df)
                    self._add_table_style(ws, df)
                
                # 自动调整列宽
                self._auto_adjust_column_width(ws)
            
            wb.save(file_path)
            
            file_size = os.path.getsize(file_path)
            total_rows = sum(len(df) for df in data_dict.values())
            logger.info(f"多工作表Excel文件导出成功: {file_path}, 大小: {file_size} bytes, 总行数: {total_rows}")
            
            return file_path
            
        except Exception as e:
            logger.error(f"多工作表Excel导出失败: {e}")
            raise
    
    def _apply_formatting(self, ws, df: pd.DataFrame):
        """应用格式化"""
        # 标题行格式
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # 边框样式
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 应用标题行格式
        for col in range(1, len(df.columns) + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # 应用数据行格式
        for row in range(2, len(df) + 2):
            for col in range(1, len(df.columns) + 1):
                cell = ws.cell(row=row, column=col)
                cell.border = thin_border
                
                # 数字格式
                if col <= len(df.columns):
                    column_name = df.columns[col - 1]
                    if df[column_name].dtype in ['int64', 'float64']:
                        cell.alignment = Alignment(horizontal="right")
                    elif df[column_name].dtype == 'datetime64[ns]':
                        cell.number_format = 'YYYY-MM-DD HH:MM:SS'
    
    def _add_table_style(self, ws, df: pd.DataFrame):
        """添加表格样式"""
        try:
            # 定义表格范围
            table_range = f"A1:{chr(64 + len(df.columns))}{len(df) + 1}"
            
            # 创建表格
            table = Table(displayName="DataTable", ref=table_range)
            
            # 设置表格样式
            style = TableStyleInfo(
                name="TableStyleMedium9",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False
            )
            table.tableStyleInfo = style
            
            # 添加表格到工作表
            ws.add_table(table)
            
        except Exception as e:
            logger.warning(f"添加表格样式失败: {e}")
    
    def _auto_adjust_column_width(self, ws):
        """自动调整列宽"""
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            # 设置列宽，最小10，最大50
            adjusted_width = min(max(max_length + 2, 10), 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def generate_filename(self, task_name: str, template: Optional[str] = None) -> str:
        """生成文件名"""
        now = datetime.now()
        
        if template and template.strip():  # 检查模板是否为空或只包含空白字符
            try:
                # 先转义模板中的百分号，避免与format()冲突
                safe_template = template.replace('%', '%%')
                
                # 支持模板变量替换
                filename = safe_template.format(
                    task_name=task_name,
                    date=now.strftime('%Y%m%d'),
                    time=now.strftime('%H%M%S'),
                    datetime=now.strftime('%Y%m%d_%H%M%S'),
                    year=now.year,
                    month=now.month,
                    day=now.day,
                    hour=now.hour,
                    minute=now.minute,
                    second=now.second,
                    # 添加更多日期格式支持
                    Y=now.year,  # 4位年份
                    y=now.strftime('%y'),  # 2位年份
                    m=now.strftime('%m'),  # 2位月份
                    d=now.strftime('%d'),  # 2位日期
                    H=now.strftime('%H'),  # 24小时制小时
                    M=now.strftime('%M'),  # 分钟
                    S=now.strftime('%S')   # 秒
                )
                
                # 恢复百分号
                filename = filename.replace('%%', '%')
                
            except (KeyError, ValueError, TypeError) as e:
                # 如果模板格式有误，使用默认格式并记录警告
                logger.warning(f"文件名模板格式错误: {template}, 错误: {e}, 使用默认格式")
                logger.warning(f"支持的模板变量: {{task_name}}, {{date}}, {{time}}, {{datetime}}, {{year}}, {{month}}, {{day}}, {{hour}}, {{minute}}, {{second}}, {{Y}}, {{y}}, {{m}}, {{d}}, {{H}}, {{M}}, {{S}}")
                filename = f"{task_name}_{now.strftime('%Y%m%d_%H%M%S')}"
        else:
            # 默认文件名格式
            filename = f"{task_name}_{now.strftime('%Y%m%d_%H%M%S')}"
        
        # 清理文件名中的非法字符
        invalid_chars = '<>:"/\\|?*'
        for char in invalid_chars:
            filename = filename.replace(char, '_')
        
        return filename
    
    def add_summary_sheet(self, wb: Workbook, summary_data: Dict[str, Any]):
        """添加汇总信息工作表"""
        ws = wb.create_sheet(title="执行汇总", index=0)
        
        # 添加汇总信息
        ws['A1'] = "执行汇总信息"
        ws['A1'].font = Font(bold=True, size=14)
        
        row = 3
        for key, value in summary_data.items():
            ws[f'A{row}'] = key
            ws[f'B{row}'] = str(value)
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
        
        # 自动调整列宽
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 30
    
    def get_file_info(self, file_path: str) -> Dict[str, Any]:
        """获取文件信息"""
        if not os.path.exists(file_path):
            return {}
        
        stat = os.stat(file_path)
        return {
            'file_path': file_path,
            'file_name': os.path.basename(file_path),
            'file_size': stat.st_size,
            'created_time': datetime.fromtimestamp(stat.st_ctime),
            'modified_time': datetime.fromtimestamp(stat.st_mtime)
        }